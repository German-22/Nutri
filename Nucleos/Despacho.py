import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import sqlite3
import os
from datetime import datetime
import openpyxl
import json
import Leer_archivo as la 
from functools import partial
ruta_txt = "/archnucl"
from csv import reader, writer
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
opciones_form = []
def leer_base():    
    global opciones_form
    try:
        conexion=sqlite3.connect(entrada_ruta_bd.get())            
        a = conexion.execute("""SELECT nombre from formulas where (sector = ? or sector = ? or nombre = ?) ORDER BY nombre;""",("Nucleos_Comasa","Nucleos_Cereales","Leche_en_Polvo_Abanderadox800g"))  
        b = a.fetchall()       
        opciones_form = b                                   
        c = conexion.execute("""SELECT destino from destinos;""")  
        d = c.fetchall() 
        e = conexion.execute("""SELECT cliente from clientes;""")  
        f = e.fetchall()   
        e = conexion.execute("""SELECT OC from ordenes where estado = ?;""",("pendiente",))  
        g = e.fetchall()  
        oc_combo_sim["values"] = list(g)  
        e = conexion.execute("""SELECT * from ordenes;""")  
        g = e.fetchall() 
        if g:
            for fila in g:
                treeoc.insert("", "end", values=fila)
        e = conexion.execute("""SELECT DISTINCT nplanilla,fecha, destino from planillas where estado != "despachado";""")  
        g = e.fetchall() 
        
        if g:
            for i in g:
                tree3.insert("", "end", values=i)         
        conexion.close()         
        
        return list(b),list(d),list(f)

    except:
        messagebox.showinfo(message="Error al Conectar con Base de Datos", title="Error de Conexion")

def selecionar_ruta(s):    
    if  s == "bd":
        ruta_guardar = []
        ruta_bd= filedialog.askopenfilename(initialdir="/", title="Seleccionar Base de Datos")                                        
        ruta_bd = ruta_bd.strip("{}")
        entrada_ruta_bd.delete("0", "end")
        entrada_ruta_bd.insert(0, str(ruta_bd))
        ruta_guardar.append(ruta_bd) 
        try: 
            archivo = open(ruta_txt + "/archivo_bd.txt", "w")
            archivo_csv = writer(archivo)
            archivo_csv.writerow(ruta_guardar)
            archivo.close()
            leer_archivo()
        except:
            messagebox.showinfo(message="Error al Configurar la Ruta", title="Ruta Erronea")
            return
    if s == "excel":
        ruta_guardar = []
        ruta_registro= filedialog.askopenfilename(initialdir="/", title="Seleccionar Plantilla Excel")                                         
        ruta_registro = ruta_registro.replace(" ", "_")
        entrada_ruta_excel.delete("0", "end")
        entrada_ruta_excel.insert(0, str(ruta_registro))
        ruta_guardar.append(ruta_registro) 
        try: 
            archivo = open(ruta_txt + "/archivo_ruta_excel.txt", "w")
            archivo_csv = writer(archivo)
            archivo_csv.writerow(ruta_guardar)
            archivo.close()
            leer_archivo()
        except:
            messagebox.showinfo(message="Error al Configurar la Ruta", title="Ruta Erronea")
            return
    if s == "planilla":
        ruta_guardar = []
        ruta_registro= filedialog.askdirectory(initialdir="/", title="Seleccionar Ruta para Guardar Excel")                                         
        ruta_registro = ruta_registro.strip("{}")
        entrada_rutae.delete("0", "end")
        entrada_rutae.insert(0, str(ruta_registro))
        ruta_guardar.append(ruta_registro) 
        try: 
            archivo = open(ruta_txt + "/archivo_ruta_planilla.txt", "w")
            archivo_csv = writer(archivo)
            archivo_csv.writerow(ruta_guardar)
            archivo.close()
            leer_archivo()
        except:
            messagebox.showinfo(message="Error al Configurar la Ruta", title="Ruta Erronea")
            return


def leer_archivo():
    global ruta_txt
    if "archnucl" in os.listdir("/"):
        ruta_txt = "/archnucl"
    else:
        try:
            os.mkdir("/archnucl")
            ruta_txt = "/archnucl"
        except:
            messagebox.showinfo(message="No se Creo el Dir. del Archivo de Rutas", title="Error al Crear el Directorio")

    if "archivo_bd.txt" in os.listdir(ruta_txt):
        archivo_bd = reader(open(ruta_txt + "/archivo_bd.txt", "r"))
        archivo_bd = (list(archivo_bd)[0])
        entrada_ruta_bd.delete("0", "end")
        entrada_ruta_bd.insert(0, (archivo_bd))
       
    else:
        messagebox.showinfo(message="Configure la Ruta a la Base de Datos", title="Ruta Erronea")
        return
    
    if "archivo_ruta_excel.txt" in os.listdir(ruta_txt):
        archivo_registro = reader(open(ruta_txt + "/archivo_ruta_excel.txt", "r"))
        archivo_registro = (list(archivo_registro)[0])
        entrada_ruta_excel.delete("0", "end")
        entrada_ruta_excel.insert(0, (archivo_registro))
        
    else:
        messagebox.showinfo(message="Configure la Ruta a La Plantilla de Excel", title="Ruta Erronea")
        return
    if "archivo_ruta_planilla.txt" in os.listdir(ruta_txt):
        archivo_registro = reader(open(ruta_txt + "/archivo_ruta_planilla.txt", "r"))
        archivo_registro = (list(archivo_registro)[0])
        entrada_rutae.delete("0", "end")
        entrada_rutae.insert(0, (archivo_registro))
        
    else:
        messagebox.showinfo(message="Configure la Ruta", title="Ruta Erronea")
        return

def cargar_datos(f):    
    stock = 0
    producto = producto_combo.get()
    for s in tree.get_children():
        tree.delete(s)
    conexion=sqlite3.connect(entrada_ruta_bd.get())  
    a = conexion.execute('''
            SELECT * FROM despacho WHERE producto=?
            ORDER BY id DESC
        ''', (producto,))
    for fila in a.fetchall():
        stock = stock + float(fila[6])
        tree.insert("", "end", values=fila) 
    e = conexion.execute("""SELECT OC from ordenes where estado = ? and producto = ?;""",("pendiente",producto))  
    g = e.fetchall()  
    oc_combo["values"] = list(g) 
    conexion.close()  
    stock_total.delete("0", "end")  
    stock_total.insert(0,stock)                    

def obtener_productos():
    conexion=sqlite3.connect(entrada_ruta_bd.get())      
    c = conexion.execute("""SELECT nombre FROM formulas where sector = ? or sector = ? ORDER BY nombre;""", ("Nucleos_Comasa","Nucleos_Cereales"))
    productos = [row[0] for row in c.fetchall()]
    conexion.close()
    return productos

def mostrar_planilla():
    planilla = tree3.item(tree3.selection())["values"][0]  
    conexion=sqlite3.connect(entrada_ruta_bd.get())  
    a = conexion.execute(f'''SELECT producto, pallet, lote, vto ,cantidad, destino, oc FROM planillas WHERE nplanilla = ? ''', (planilla,))
    b = a.fetchall()
    for s in tree4.get_children():
            tree4.delete(s)
    for fila in b:
            tree4.insert("", "end", values=fila)
    conexion.close()    
    return

def eliminar_planilla():
    planilla = tree3.item(tree3.selection())["values"][0] 
    
    conexion=sqlite3.connect(entrada_ruta_bd.get()) 
    try:
        a = conexion.execute(
            'SELECT cantidad, id_pallet, oc FROM planillas WHERE nplanilla = ?', 
            (planilla,)
        )
        b = a.fetchall()
        for oc in b:
            conexion.execute(
            "UPDATE despacho SET cantidad = cantidad + ? WHERE id = ?", (oc[0],oc[1])
            
        )
        conexion.commit()
        for oc in b:
            conexion.execute(
                "UPDATE ordenes SET pendiente = pendiente + ? WHERE OC = ?", 
                (oc[0],oc[2] )
            )
        conexion.commit()

    except Exception as e:
        conexion.rollback()
        print("Ocurrió un error al actualizar la base de datos:", e)
        conexion.close()
        return
    conexion.execute("DELETE FROM planillas WHERE nplanilla = ?", (planilla,))
    conexion.commit()  
    conexion.close()  
    seleccionado = tree3.selection()
    if not seleccionado:
        messagebox.showwarning("Sin selección", "Selecciona un registro para eliminar.")
        return
    tree3.delete(seleccionado[0])  
    cargar_datos(planilla)     
    return

def agregar_sim():   
    oc = oc_combo_sim.get()
    producto = producto_combo_sim.get()    
    seleccion = parsear_pallets(pallets_entry_sim.get().strip())
    lote = entry_lote_sim.get()
    vto = entry_vto_sim.get()
    #responsable = resp_sim.get()
    cajas = entry_cajas_sim.get()
    if oc =="":
        messagebox.showwarning("Error","Seleccione un OC")
        return
    if producto =="" or pallets_entry_sim.get()=="" or lote == "" or vto == "" or cajas == "":
        messagebox.showwarning("Error","Complete los Campos")
        return
    conexion=sqlite3.connect(entrada_ruta_bd.get()) 
    a = conexion.execute(f'''SELECT destino FROM ordenes WHERE OC = ?''', (oc,))
    b = a.fetchall()
    conexion.close() 
    for i in seleccion:
        tree_sim.insert("", "end", values=["",producto,i,lote,vto,cajas,b[0][0],oc])
    return

def agregar_cajas_sim():
    cajas = cajas_entry_sim.get()
    producto = producto_combo_sim.get()
    oc = oc_combo_sim.get()
    lote = entry_lote_sim.get()
    vto = entry_vto_sim.get()
    #responsable = resp_sim.get()
    pallet = pallet_sim.get()
    if oc =="":
        messagebox.showwarning("Error","Seleccione un OC")
        return
    if pallet == "" or producto =="" or lote == "" or vto == "" or cajas == "":
        messagebox.showwarning("Error","Complete los Campos")
        return
    conexion=sqlite3.connect(entrada_ruta_bd.get()) 
    a = conexion.execute(f'''SELECT destino FROM ordenes WHERE OC = ?''', (oc,))
    b = a.fetchall()
    conexion.close()
    tree_sim.insert("", "end", values=["",producto,pallet,lote,vto,cajas,b[0][0],oc])
    return
   

def agregar():
    
    producto = producto_combo.get()    
    seleccion = parsear_pallets(pallets_entry.get().strip())
    
    if seleccion == "":
        messagebox.showwarning("Error","Error en valor ingresado")
        return
    oc = oc_combo.get()
    try:
        conexion=sqlite3.connect(entrada_ruta_bd.get())         
        a =conexion.execute(f'''SELECT producto FROM ordenes WHERE oc = ?''', (oc,))
        d = a.fetchall()
        if d[0][0]!=producto:
            messagebox.showwarning("Error","La OC no corresponde con el Producto")
            conexion.close()
            return       
        condiciones = "producto = ? AND (" + " OR ".join(["pallet = ?"] * len(seleccion)) + ") and estado = ? and cantidad != ?"
        valores = [producto] + seleccion + ["liberado", 0]
        f = conexion.execute(f'''SELECT id ,producto, pallet, lote, vto ,cantidad FROM despacho WHERE {condiciones}''', valores)
        d = f.fetchall()
        b = conexion.execute(f'''SELECT destino FROM ordenes WHERE OC = ?''', (oc,))
        a = b.fetchall()
        
        for fila in d:    
            duplicado = False
            for item in tree2.get_children():
                valores_existentes = tree2.item(item, "values")
                if (
                    str(valores_existentes[1]) == str(fila[1]) and  # producto
                    str(valores_existentes[2]) == str(fila[2]) and  # pallet
                    str(valores_existentes[3]) == str(fila[3])      # lote
                ):
                    duplicado = True
                    break

            if not duplicado:
                tree2.insert("", "end", values=list(fila) + [a[0][0]] + [oc])      

    except Exception as e:
        conexion.rollback()
        messagebox.showerror("Error", str(e))


def agregar_oc():
    
    producto = producto_combo_oc.get()    
    oc = noc.get()
    cantidad = entry_cantidad.get()
    destino = combo_destino.get()
    cliente = combo_cliente.get()
    fecha = fecha_pedido.get()
    if not oc or not destino or not cantidad or not cliente or not fecha:
            messagebox.showwarning("Datos incompletos", "Completa los campos")
            return

    try:
        cantidad = int(cantidad)
        if cantidad <= 0:
            raise ValueError
    except ValueError:
        messagebox.showwarning("Dato inválido", "Cantidad de cajas debe ser un número entero positivo.")
        return
    
    conexion=sqlite3.connect(entrada_ruta_bd.get())   
    try:
        conexion.execute('''
            INSERT INTO ordenes (OC,Producto,Cajas,pendiente,Destino,Cliente,FechaEntrega,Estado)
            VALUES (?, ?, ?, ?, ?, ?,?,?)
        ''', (oc,producto, cantidad,cantidad,destino,cliente,fecha,"pendiente"))
        conexion.commit()
        
        c = conexion.execute(f'''SELECT id,OC,Producto,Cajas,pendiente,Destino,Cliente,FechaEntrega,FechaDespacho,Estado FROM ordenes''')
        filas = c.fetchall()
        for s in treeoc.get_children():
            treeoc.delete(s)
        for fila in filas:
            treeoc.insert("", "end", values=fila)
        conexion.close()
    except:
        messagebox.showinfo(message="Dato Invalido", title="Error")
        conexion.close()
    return

def finalizar_oc():
    id = treeoc.item(treeoc.selection())["values"][0]      
    conexion=sqlite3.connect(entrada_ruta_bd.get())    
    conexion.execute("""UPDATE ordenes SET estado = "finalizado" WHERE id = ?;""" ,(id,))
    conexion.commit()
    for s in treeoc.get_children():
        treeoc.delete(s)
    e = conexion.execute("""SELECT * from ordenes;""")  
    g = e.fetchall() 
    if g:
        for fila in g:
            treeoc.insert("", "end", values=fila)
    conexion.close()

    return
    

def agregar_cajas():
    cajas = cajas_entry.get()
    producto = producto_combo.get()
    oc = oc_combo.get()
    id = tree.item(tree.selection())["values"][0]  
    if id == "":
          messagebox.showinfo(message="Seleccione un Pallet", title="Error")
          return
    if cajas == "" or float(cajas) == False:
        messagebox.showinfo(message="Error en Valor Ingresado", title="Error")
        return

    conexion=sqlite3.connect(entrada_ruta_bd.get()) 
    a =conexion.execute(f'''SELECT producto FROM ordenes WHERE oc = ?''', (oc,))
    d = a.fetchall()
    if d[0][0]!=producto:
        messagebox.showwarning("Error","La OC no corresponde con el Producto")
        conexion.close()
        return    
    a = conexion.execute(f'''SELECT id ,producto, pallet, lote, vto ,cantidad FROM despacho WHERE id = ? ''', (id,))
    b = a.fetchall()
    c = list(b[0])
    if int(c[5])>=int(cajas):
        c[5]=cajas    

        x = conexion.execute(f'''SELECT destino FROM ordenes WHERE OC = ?''', (oc,))
        y = x.fetchall()    
    # Revisar si ya existe en tree2 una fila con mismo producto, pallet y lote
        duplicado = False
        for item in tree2.get_children():
            valores_existentes = tree2.item(item, "values")
            if (
                str(valores_existentes[1]) == str(c[1]) and  # producto
                str(valores_existentes[2]) == str(c[2]) and  # pallet
                str(valores_existentes[3]) == str(c[3])      # lote
            ):
                duplicado = True
                break

        if not duplicado:
            tree2.insert("", "end", values=list(c) + [y[0][0]] + [oc]) 
    else:
        messagebox.showinfo(message="Este Pallet no Tiene Suficientes Cajas", title="Error")
    return

def parsear_pallets(entrada):
    resultado = []
    for parte in entrada.split(','):
        if '-' in parte:
            inicio, fin = parte.split('-')
            resultado.extend(range(int(inicio), int(fin)+1))
        else:
            resultado.append(int(parte))
    return resultado

def simulacion():
    valores = []  
       
    for item_id in tree_sim.get_children():
        valores.append(tree_sim.item(item_id)["values"])
    if not valores:
        messagebox.showwarning("Advertencia", "No hay datos para exportar.")
        return
    crear_planilla(valores)

def crear_planilla(valores):
    try:  
        plantilla_path = entrada_ruta_excel.get()
        plantilla_path = plantilla_path.replace("_", " ")
        carpeta = entrada_rutae.get()
        # Diccionario para acumular resultados
        conteo = defaultdict(int)
        # Contar cajas por producto y lote}
        
        for _,producto,_, lote,_,cajas,_,_ in valores:
            clave = (producto, lote)            
            conteo[clave] += float(cajas)
        
        
        conteo2 = defaultdict(int)
        conteo3 = defaultdict(int)
        conteo4 = defaultdict(int)
        # Recorremos la lista y contamos
        for _,producto,_, lote,_,cajas,_,_ in valores:
            conteo2[(producto)] += 1
        
        lotes_por_producto = {}

        for _,producto,_, lote,vto,cajas,_,_ in valores:
            clave = (producto, lote)
            conteo4[clave] = vto
            if producto not in lotes_por_producto:
                lotes_por_producto[producto] = set()                
                conteo3[clave] += float(cajas)            
            lotes_por_producto[producto].add(lote)

        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active  
        fila_inicio = 11
        columna_inicio = 2   
            
         
        for producto,_ in conteo2.items():              
            repeticiones = conteo2[(producto)]               
            ws.cell(row=fila_inicio, column=columna_inicio, value=producto)      
            n_columnas = 2
            n_filas = int(repeticiones)       

            # Calcular celda final
            columna_fin = columna_inicio + n_columnas - 1
            fila_fin = fila_inicio + n_filas - 1

            # Convertir columnas a letras
            col_letra_inicio = get_column_letter(columna_inicio)
            col_letra_fin = get_column_letter(columna_fin)

            # Rango a combinar
            rango = f"{col_letra_inicio}{fila_inicio}:{col_letra_fin}{fila_fin}"

            # Combinar celdas
            ws.merge_cells(rango)               

            # (Opcional) centrar el texto
            
            ws.cell(row=fila_inicio, column=2).alignment = Alignment(horizontal="center", vertical="center")
            fila_inicio = fila_inicio + n_filas

        fila_inicio = 45
        columna_inicio = 2 
        for producto, lotes in lotes_por_producto.items():              
            repeticiones = len(lotes)              
            ws.cell(row=fila_inicio, column=columna_inicio, value=producto)      
            n_columnas = int(repeticiones) 
            n_filas = 1
            c = columna_inicio       
            for i in lotes:
                ws.cell(row=46, column=c, value=i)  
                ws.cell(row=47, column=c, value=conteo[producto,i])     
                ws.cell(row=48, column=c, value=str(conteo4[producto,i]))                        
                c+=1
            
            # Calcular celda final
            columna_fin = columna_inicio + n_columnas - 1
            fila_fin = fila_inicio + n_filas - 1

            # Convertir columnas a letras
            col_letra_inicio = get_column_letter(columna_inicio)
            col_letra_fin = get_column_letter(columna_fin)

            # Rango a combinar
            rango = f"{col_letra_inicio}{fila_inicio}:{col_letra_fin}{fila_fin}"

            # Combinar celdas
            ws.merge_cells(rango)               

            
            ws.cell(row=fila_inicio, column=2).alignment = Alignment(horizontal="center", vertical="center")
            fila_inicio = fila_inicio + n_filas

        fila_inicio = 10
        pallet_anterior = 0
        for v in valores:
            
            if pallet_anterior == v[2]:
                ws.cell(row=fila_inicio, column=5+2,value=v[3])
                ws.cell(row=fila_inicio, column=6+2,value=v[5])                
            else:
                fila_inicio+=1                
                ws.cell(row=fila_inicio, column=4,value=v[2])
                ws.cell(row=fila_inicio, column=5,value=v[3])
                ws.cell(row=fila_inicio, column=6,value=v[5])                
                pallet_anterior = v[2]


        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        archivo_salida = os.path.join(carpeta, f"{valores[0][1]}_{fecha_actual}.xlsx")
        wb.save(archivo_salida)
        messagebox.showinfo("Éxito", f"Archivo exportado: {archivo_salida}")

    except Exception as e:
        messagebox.showerror("Error", str(e))
        return

def ejecutar_exportacion():    
    valores = []  
    hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conteo5 = defaultdict(float)
   
    for item_id in tree2.get_children():
        valores.append(tree2.item(item_id)["values"])
    if not valores:
        messagebox.showwarning("Advertencia", "No hay datos para exportar.")
        return
    for fila in valores:
        try:
            cajas = float(fila[5])
            oc = fila[7]
            conteo5[oc] += cajas
        except (IndexError, ValueError, TypeError) as e:
            print(f"Error procesando fila {fila}: {e}")
    try:
        conexion = sqlite3.connect(entrada_ruta_bd.get())
        for oc, total_cajas in conteo5.items():
                a =conexion.execute(f'''SELECT pendiente FROM ordenes WHERE oc = ?''', (oc,))
                d = a.fetchall()         
            
                if total_cajas>d[0][0]:
                    messagebox.showwarning("Error","La cantidad de Cajas es Mayor al Pendiente de Despacho")
                    conexion.close()
                    return     
        
        for v in valores:
            
            cursor = conexion.execute('SELECT cantidad FROM despacho WHERE id = ?', (v[0],))
            resultado = cursor.fetchone()

            if resultado is not None:
                cantidad_actual = resultado[0]
                
                if float(cantidad_actual) >= float(v[5]):
                    
                    conexion.execute('''
                        UPDATE despacho 
                        SET cantidad = (cantidad - ?)            
                        WHERE id = ?
                    ''', (v[5], v[0]))
                else:
                    messagebox.showwarning(
                        "Cantidad insuficiente",
                        f"No se puede descontar {v[5]} del ID {v[0]}: solo hay {cantidad_actual} disponible."
                    )
                    conexion.rollback()
                    conexion.close()
                    return
            else:
                messagebox.showerror("Error", f"No se encontró el ID {v[0]} en la tabla despacho.")
                conexion.rollback()
                return
        conexion.commit()
        for oc, total_cajas in conteo5.items():
            conexion.execute(
                "UPDATE ordenes SET pendiente = pendiente - ? WHERE OC = ?", 
                (total_cajas, oc)
            )
        conexion.commit()
            
    except Exception as e:
        conexion.rollback()
        print("Error al actualizar la base de datos:", e)
        conexion.close()

    # Obtener datos del treeview         
    try:
        conexion = sqlite3.connect(entrada_ruta_bd.get())
        # Obtener destino desde ordenes
        resultado = conexion.execute('SELECT Destino FROM ordenes WHERE OC = ?', (oc_combo.get(),)).fetchone()
        if not resultado:
            messagebox.showerror("Error", "No se encontró el destino para la OC especificada.")
            conexion.close()
            return
        destino = resultado[0]
        
        # Crear nombre de planilla
        nplanilla = str(valores[0][1]) + hora
        
        # Insertar en planillas
        for v in valores:
            conexion.execute('''
                INSERT INTO planillas (producto,nplanilla, fecha,pallet,cantidad,vto,lote,destino,oc,id_pallet)
                VALUES (?, ?,?, ?,?,?,?,?,?,?)
            ''', (v[1],nplanilla, hora,v[2],v[5],v[4],v[3],destino,v[7],v[0]))
        conexion.commit()

        # Mostrar la nueva planilla en tree3
        planilla_row = conexion.execute('SELECT DISTINCT nplanilla,fecha,destino FROM planillas WHERE nplanilla = ?', (nplanilla,)).fetchone()
        
        if planilla_row:
            tree3.insert("", "end", values=planilla_row)
    
        
        conexion.close()
        crear_planilla(valores) 
        cargar_datos(v)

    except Exception as e:
        conexion.rollback()
        conexion.close()
        messagebox.showerror("Error", f"Ocurrió un error:\n{e}")
        return
    
def limpiar_planilla():
    for s in tree2.get_children():
        tree2.delete(s)
    return

def validar_entrada2(texto):
    return all(c.isdigit() for c in texto)    

def validar_entrada(texto):
    if texto == "":
        return True  # permitir borrar todo
    if texto.startswith('-'):
        return False
    if '--' in texto:
        return False
    return all(c.isdigit() or c == '-' for c in texto)

def actualizar_pallet():
    pallet = n_pallet.get()
    lote = entry_lote.get()
    vtoc = entry_vto.get()
    comen = comentario.get()    
    cajas = cajas_act.get() 
    conexion=sqlite3.connect(entrada_ruta_bd.get())
    lista2 = tree.item(tree.selection())["values"]   
    id = lista2[0]
    if pallet!="":   
        lista2[3] = pallet
        conexion.execute("""UPDATE despacho SET pallet = ? WHERE id = ?;""" ,(pallet,id))
        conexion.commit()
    if lote!="":
        lista2[4] = lote
        conexion.execute("""UPDATE despacho SET lote = ? WHERE id = ?;""" ,(lote,id))
        conexion.commit()
    if vtoc!="":
        lista2[5] = lote
        if len(vtoc)==10:
            dia = vtoc[0:2]
            mes = vtoc[3:5]
            año = vtoc[6:10]        
            rev = (año) + "-" + (mes) + "-" + (dia)                  
            try:            
                datetime.strptime(rev, "%Y-%m-%d")            
            except:            
                messagebox.showinfo(message="Error en Fecha de Vto", title="Error de Fecha")
                return       
        conexion.execute("""UPDATE despacho SET vto = ? WHERE id = ?;""" ,(vtoc,id))
        conexion.commit()
    if cajas!="":
        lista2[6] = lote
        conexion.execute("""UPDATE despacho SET cantidad = ? WHERE id = ?;""" ,(cajas,id))
        conexion.commit()
    if comen!="":
        lista2[8] = lote
        conexion.execute("""UPDATE despacho SET comentario = ? WHERE id = ?;""" ,(comen,id))
        conexion.commit()
    conexion.close()
    cargar_datos(lote)
    return

def cargar_pallet():
    pallet = n_pallet.get()
    lote = entry_lote.get()
    vto = entry_vto.get()
    comen = comentario.get()    
    cajas = cajas_act.get() 
    responsable = resp.get()
   
    if cajas == "" or vto=="" or lote == "" or pallet == "" or responsable =="":
        messagebox.showinfo(message="Ingrese los Datos", title="Error")
        return
    producto = producto_combo.get()
    if producto == "":
        messagebox.showinfo(message="Seleccione el Producto", title="Error")
        return
    
    hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conexion=sqlite3.connect(entrada_ruta_bd.get())
    conexion.execute('''
                INSERT INTO despacho (hora,producto,pallet, lote,vto,cantidad, comentario, responsable)
                VALUES (?, ?, ?, ?, ?, ?,?,?)
            ''', (hora,producto, pallet, lote,vto ,cajas, comen,responsable))
    conexion.commit()
    conexion.close
    cargar_datos(hora)
    return

def eliminar_pallet():
    lista2 = tree.item(tree.selection())["values"]   
    id = lista2[0]
    conexion=sqlite3.connect(entrada_ruta_bd.get())
    conexion.execute("DELETE FROM despacho WHERE id = ?", (id,))
    conexion.close()
    cargar_datos(id)
    return
       

def filtrar_opciones(formula,opciones,s):    
    if opciones == "form_s":
        opcion = opciones_form
        ent = combo_var.get()
        entrada = combo_var.get().lower()
    elif opciones == "form":
        opcion = opciones_form
        entrada = combo_var2.get().lower()
        ent = combo_var2.get()
    elif opciones == "oc":
        opcion = opciones_form
        entrada = combo_var3.get().lower()
        ent = combo_var3.get()
        
    # Filtrar opciones que contengan el texto
    filtradas = [op for op in opcion if entrada in op[0].lower()]
    
    # Guardar posición del cursor y texto actual
    cursor_pos = formula.index(tk.INSERT)
    
    # Actualizar valores del Combobox
    formula['values'] = filtradas if filtradas else opcion
    
    # Restaurar el texto y la posición del cursor
    formula.delete(0, tk.END)
    formula.insert(0, ent)
    formula.icursor(cursor_pos)
    
    # Autocompletar si hay una sola opción
    if len(filtradas) == 1:
        formula.delete(0, tk.END)
        formula.insert(0, filtradas[0])
        formula.icursor(tk.END)

    # Mostrar menú desplegable
    formula.event_generate('<Down>')
# ---------------- Interfaz ----------------

def ordenar(col):    
    if col == "pallet":
        elem = []
        elem_ordenado = []
        for t in tree.get_children():
            elem.append((tree.item(t)))       
        elem_ordenado = sorted(elem, key=lambda x: x["values"][3])        
        for s in tree.get_children():
            tree.delete(s)
        for s in elem_ordenado:
            tree.insert("", tk.END,values=(s["values"]))

root = tk.Tk()
root.title("Despacho")
root.geometry("2000x800")

# Crear contenedor de pestañas
notebook = ttk.Notebook(root,width=2000, height=800)
notebook.place(x=0, y=0, relheight=1, relwidth=1)

# Pestaña: Exportar Producción
frame_exportar = ttk.Frame(notebook)
notebook.add(frame_exportar, text="Despacho")

frame_sim_desp = ttk.Frame(notebook)
notebook.add(frame_sim_desp, text="Simular Despacho")
# Contenido de la pestaña Simular despacho
ttk.Label(frame_sim_desp, text="Producto:").place(relx=0,rely=0.01)
combo_var = tk.StringVar()
combo_var2 = tk.StringVar()
producto_combo_sim = ttk.Combobox(frame_sim_desp,width=50,textvariable=combo_var)
producto_combo_sim.place(relx=0.05,rely=0.01)
producto_combo_sim.bind('<Return>', partial(filtrar_opciones,producto_combo_sim,"form_s"))
ttk.Label(frame_sim_desp, text="OC:").place(relx=0.02,rely=0.08)
oc_combo_sim = ttk.Combobox(frame_sim_desp, width=50)
oc_combo_sim.place(relx=0.05,rely=0.08)

contenedor_tree_sim = ttk.Frame(frame_sim_desp)
contenedor_tree_sim.place(relx=0.01, rely=0.18, relwidth=0.6, relheight=0.4)  # Ajusta si es necesario
ttk.Label(contenedor_tree_sim, text="Produccion y Despacho:").place(relx=0,rely=0)
# Scrollbars
scrollbar_y = ttk.Scrollbar(contenedor_tree_sim, orient="vertical")
scrollbar_x = ttk.Scrollbar(contenedor_tree_sim, orient="horizontal")

# Treeview con scrollbars
tree_sim = ttk.Treeview(
    contenedor_tree_sim,
    columns=("","Producto", "Pallet", "Lote", "Vto", "Cajas","Destino", "OC"),
    show='headings',
    yscrollcommand=scrollbar_y.set,
    xscrollcommand=scrollbar_x.set,
    height=10    
)

scrollbar_y.config(command=tree_sim.yview)
scrollbar_x.config(command=tree_sim.xview)

# Configurar columnas
for col in tree_sim["columns"]:
    tree_sim.heading(col, text=col)
    if col == "":
        tree_sim.column(col, width=0, stretch=False)
    elif col in ("Vto", "Cajas", "Pallet"):        
        tree_sim.column(col, anchor="center", width=70,stretch=False)
    elif col == "Producto":
        tree_sim.column(col, anchor="center", width=200,stretch=False)
    else:
        
        tree_sim.column(col, anchor="center", width=100,stretch=False)

# Empaquetar widgets
tree_sim.place(relx=0,rely=0.06, relheight=0.85,relwidth=0.96)
scrollbar_y.place(relx=0.95,rely=0.06, relheight=0.85,relwidth=0.05)
scrollbar_x.place(relx=0,rely=0.9, relheight=0.1,relwidth=0.98)


ttk.Label(frame_sim_desp, text="Pallets (ej: 1,3,5-8):").place(relx=0.3,rely=0)

pallet_frame_sim = ttk.Frame(frame_sim_desp)
pallet_frame_sim.place(relx=0.3,rely=0.03)

vcmd = (root.register(validar_entrada), '%P')
pallets_entry_sim = ttk.Entry(pallet_frame_sim, width=30, validate="key", validatecommand=vcmd)
pallets_entry_sim.pack(side="left", padx=(5, 0)) 

vcmd = (root.register(validar_entrada2), '%P')
cajas_frame_sim = ttk.Frame(frame_sim_desp)
cajas_frame_sim.place(relx=0.3,rely=0.09,relheight=0.1,relwidth=0.2)
ttk.Label(cajas_frame_sim,text="N° de cajas:").place(relx=0,rely=0)
ttk.Label(cajas_frame_sim,text="N° de Pallet:").place(relx=0,rely=0.5)
cajas_entry_sim = ttk.Entry(cajas_frame_sim, width=10,validate="key", validatecommand=vcmd)
cajas_entry_sim.place(relx=0.3,rely=0)
pallet_sim = ttk.Entry(cajas_frame_sim, width=10,validate="key", validatecommand=vcmd)
pallet_sim.place(relx=0.3,rely=0.5)
ttk.Button(cajas_frame_sim, text="Agregar a Planilla", command=agregar_cajas_sim).place(relx=0.6,rely=0.2)
ttk.Button(pallet_frame_sim, text="Agregar a Planilla", command=agregar_sim).pack(side="left", padx=(5, 0))  

ttk.Label(frame_sim_desp, text="Lote:").place(relx=0.5,rely=0.01)
entry_lote_sim = ttk.Entry(frame_sim_desp, width=30)
entry_lote_sim.place(relx=0.55,rely=0.01)
ttk.Label(frame_sim_desp, text="Vto:").place(relx=0.5,rely=0.05)
entry_vto_sim = ttk.Entry(frame_sim_desp, width=30)
entry_vto_sim.place(relx=0.55,rely=0.05)
ttk.Label(frame_sim_desp, text="Cajas/pallet:").place(relx=0.5,rely=0.09)

entry_cajas_sim = ttk.Entry(frame_sim_desp, width=30,validate="key", validatecommand=vcmd)
entry_cajas_sim.place(relx=0.55,rely=0.09) 
#ttk.Label(frame_sim_desp, text="Responsable:").place(relx=0.5,rely=0.13)
#resp_sim = ttk.Entry(frame_sim_desp, width=30)
#resp_sim.place(relx=0.55,rely=0.13)
btn_exp_sim = ttk.Button(frame_sim_desp, text="Generar Planilla",command = simulacion)
btn_exp_sim.place(relx=0.69, rely=0.35,relheight=0.08)

# Contenido de la pestaña Despacho
ttk.Label(frame_exportar, text="Producto:").place(relx=0,rely=0.01)
producto_combo = ttk.Combobox(frame_exportar,width=50,textvariable=combo_var2)
producto_combo.place(relx=0.05,rely=0.01)
producto_combo.bind('<Return>', partial(filtrar_opciones,producto_combo,"form"))
ttk.Label(frame_exportar, text="OC:").place(relx=0.02,rely=0.13)
oc_combo = ttk.Combobox(frame_exportar, width=50)
oc_combo.place(relx=0.05,rely=0.13) 


ttk.Label(frame_exportar, text="Stock Total:").place(relx=0,rely=0.07)
stock_total = ttk.Entry(frame_exportar, width=20)
stock_total.place(relx=0.05,rely=0.07) 

ttk.Label(frame_exportar, text="Pallets (ej: 1,3,5-8):").place(relx=0.65,rely=0.3)

pallet_frame = ttk.Frame(frame_exportar)
pallet_frame.place(relx=0.65,rely=0.35)

vcmd = (root.register(validar_entrada), '%P')
pallets_entry = ttk.Entry(pallet_frame, width=30, validate="key", validatecommand=vcmd)
pallets_entry.pack(side="left", padx=(5, 0)) 

ttk.Button(pallet_frame, text="Agregar a Planilla", command=agregar).pack(side="left", padx=(5, 0))  
ttk.Label(frame_exportar, text="Cantidad de Cajas:").place(relx=0.3,rely=0.05)
vcmd = (root.register(validar_entrada), '%P')
cajas_act = ttk.Entry(frame_exportar, width=30,validate="key", validatecommand=vcmd)
cajas_act.place(relx=0.37,rely=0.05) 
ttk.Label(frame_exportar, text="N° de Pallet:").place(relx=0.3,rely=0.01)
n_pallet = ttk.Entry(frame_exportar, width=30,validate="key", validatecommand=vcmd)
n_pallet.place(relx=0.37,rely=0.01) 
ttk.Label(frame_exportar, text="Lote:").place(relx=0.3,rely=0.09)
entry_lote = ttk.Entry(frame_exportar, width=30)
entry_lote.place(relx=0.37,rely=0.09)
ttk.Label(frame_exportar, text="Vto:").place(relx=0.3,rely=0.13)
entry_vto = ttk.Entry(frame_exportar, width=30)
entry_vto.place(relx=0.37,rely=0.13)
ttk.Label(frame_exportar, text="Comentario:").place(relx=0.5,rely=0.05)
comentario = ttk.Entry(frame_exportar, width=30)
comentario.place(relx=0.55,rely=0.05) 
ttk.Label(frame_exportar, text="Responsable:").place(relx=0.5,rely=0.01)
resp = ttk.Entry(frame_exportar, width=30)
resp.place(relx=0.55,rely=0.01) 

ttk.Button(frame_exportar, text="Actualizar Pallet", command=actualizar_pallet).place(relx=0.7,rely=0.01)
ttk.Button(frame_exportar, text="Cargar Pallet", command=cargar_pallet).place(relx=0.7,rely=0.05)
ttk.Button(frame_exportar, text="Eliminar Pallet", command=eliminar_pallet).place(relx=0.7,rely=0.09) 
# 🟦 Nueva sección: N° de cajas + botón agregar_cajas
cajas_frame = ttk.Frame(frame_exportar)
cajas_frame.place(relx=0.65,rely=0.4)

ttk.Label(cajas_frame, text="N° de cajas:").pack(side="left", padx=(0, 5))
vcmd = (root.register(validar_entrada2), '%P')
cajas_entry = ttk.Entry(cajas_frame, width=10,validate="key", validatecommand=vcmd)
cajas_entry.pack(side="left", padx=(0, 5))

ttk.Button(cajas_frame, text="Agregar a Planilla", command=agregar_cajas).pack(side="left")

# Tabla de resultados
# Contenedor para tree + scrollbars
contenedor_tree = ttk.Frame(frame_exportar)
contenedor_tree.place(relx=0.01, rely=0.18, relwidth=0.6, relheight=0.4)  # Ajusta si es necesario
ttk.Label(contenedor_tree, text="Produccion y Despacho:").place(relx=0,rely=0)
# Scrollbars
scrollbar_y = ttk.Scrollbar(contenedor_tree, orient="vertical")
scrollbar_x = ttk.Scrollbar(contenedor_tree, orient="horizontal")

# Treeview con scrollbars
tree = ttk.Treeview(
    contenedor_tree,
    columns=("id", "Fecha", "Producto", "Pallet", "Lote", "Vto", "Cajas","Estado", "Comentario","Responsable"),
    show='headings',
    yscrollcommand=scrollbar_y.set,
    xscrollcommand=scrollbar_x.set,
    height=10    
)

scrollbar_y.config(command=tree.yview)
scrollbar_x.config(command=tree.xview)

# Configurar columnas
for col in tree["columns"]:
    tree.heading(col, text=col)
    if col == "id":
        tree.column(col, width=0, stretch=False)
    elif col in ("Pallet"):
        tree.column(col, anchor="center", width=70,stretch=False)
        tree.heading(col, text=col,command=lambda: ordenar("pallet"))
    elif col in ("Vto", "Cajas"):
        tree.column(col, anchor="center", width=70,stretch=False)
    elif col == "Producto":
        tree.column(col, anchor="center", width=200,stretch=False)
    else:
        tree.column(col, anchor="center", width=100,stretch=False)

# Empaquetar widgets
tree.place(relx=0,rely=0.06, relheight=0.85,relwidth=0.96)
scrollbar_y.place(relx=0.95,rely=0.06, relheight=0.85,relwidth=0.05)
scrollbar_x.place(relx=0,rely=0.9, relheight=0.1,relwidth=0.98)

contenedor_tree2 = ttk.Frame(frame_exportar)
contenedor_tree2.place(relx=0.01, rely=0.6, relwidth=0.3, relheight=0.3)  # Ajusta si es necesario
ttk.Label(contenedor_tree2, text="Planilla Actual:").place(relx=0,rely=0)
# Scrollbars
scrollbar_y = ttk.Scrollbar(contenedor_tree2, orient="vertical")
scrollbar_x = ttk.Scrollbar(contenedor_tree2, orient="horizontal")

# Treeview con scrollbars
tree2 = ttk.Treeview(
    contenedor_tree2,
    columns=("id","Producto", "Pallet", "Lote","Vto" ,"Cajas","Destino","OC"), show='headings', 
    yscrollcommand=scrollbar_y.set,
    xscrollcommand=scrollbar_x.set,
    height=10
)

scrollbar_y.config(command=tree2.yview)
scrollbar_x.config(command=tree2.xview)

# Configurar columnas
for col in tree2["columns"]:
    tree2.heading(col, text=col)
    if col == "id":
        tree2.column(col, width=0, stretch=False)
    elif col in ("Pallet", "Vto", "Cajas"):
        tree2.column(col, anchor="center", width=70)
    elif col == "Producto":
        tree2.column(col, anchor="center", width=200)
    else:
        tree2.column(col, anchor="center", width=100)

# Empaquetar widgets
tree2.place(relx=0,rely=0.08, relheight=0.85,relwidth=0.96)
scrollbar_y.place(relx=0.95,rely=0.08, relheight=0.85,relwidth=0.05)
scrollbar_x.place(relx=0,rely=0.9, relheight=0.1,relwidth=0.98)
ttk.Button(frame_exportar, text="Exportar", command=ejecutar_exportacion).place(relx=0.31,rely=0.7)
#Planillas Activas

contenedor_tree3 = ttk.Frame(frame_exportar)
contenedor_tree3.place(relx=0.38, rely=0.6, relwidth=0.25, relheight=0.3)  # Ajusta si es necesario
ttk.Button(frame_exportar, text="Exportar", command=ejecutar_exportacion).place(relx=0.31,rely=0.7)
ttk.Button(frame_exportar, text="Limpiar Planilla", command=limpiar_planilla).place(relx=0.31,rely=0.75)
ttk.Button(frame_exportar, text="Mostrar", command=mostrar_planilla).place(relx=0.63,rely=0.7)
ttk.Button(frame_exportar, text="Eliminar", command=eliminar_planilla).place(relx=0.63,rely=0.75)
ttk.Label(contenedor_tree3, text="Planillas Activas:").place(relx=0,rely=0)
# Scrollbars
scrollbar_y = ttk.Scrollbar(contenedor_tree3, orient="vertical")
scrollbar_x = ttk.Scrollbar(contenedor_tree3, orient="horizontal")

# Treeview con scrollbars
tree3 = ttk.Treeview(
    contenedor_tree3,
    columns=("N° de Planilla","Fecha","Destino"), show='headings', 
    yscrollcommand=scrollbar_y.set,
    xscrollcommand=scrollbar_x.set,
    height=10
)

scrollbar_y.config(command=tree3.yview)
scrollbar_x.config(command=tree3.xview)

# Configurar columnas
for col in tree3["columns"]:
    tree3.heading(col, text=col)
    if col == "id" or col == "N° de Planilla":
        tree3.column(col, width=0, stretch=False)
    elif col in ("Pallet", "Vto", "Cajas"):
        tree3.column(col, anchor="center", width=70)
    elif col == "Producto":
        tree3.column(col, anchor="center", width=200)
    else:
        tree3.column(col, anchor="center", width=100)

# Empaquetar widgets
tree3.place(relx=0,rely=0.08, relheight=0.85,relwidth=0.96)
scrollbar_y.place(relx=0.95,rely=0.08, relheight=0.85,relwidth=0.05)
scrollbar_x.place(relx=0,rely=0.9, relheight=0.1,relwidth=0.98)

#Mostrar contenido de planillas Activas

contenedor_tree4 = ttk.Frame(frame_exportar)
contenedor_tree4.place(relx=0.7, rely=0.6, relwidth=0.25, relheight=0.3)  # Ajusta si es necesario
ttk.Label(contenedor_tree4, text="Contenido de Planillas Activas:").place(relx=0,rely=0)
# Scrollbars
scrollbar_y = ttk.Scrollbar(contenedor_tree4, orient="vertical")
scrollbar_x = ttk.Scrollbar(contenedor_tree4, orient="horizontal")

# Treeview con scrollbars
tree4 = ttk.Treeview(
    contenedor_tree4,
    columns=("Producto", "Pallet", "Lote","Vto" ,"Cajas","Destino","OC"), show='headings', 
    yscrollcommand=scrollbar_y.set,
    xscrollcommand=scrollbar_x.set,
    height=10
)

scrollbar_y.config(command=tree4.yview)
scrollbar_x.config(command=tree4.xview)

# Configurar columnas
for col in tree4["columns"]:
    tree4.heading(col, text=col)
    if col == "id":
        tree4.column(col, width=0, stretch=False)
    elif col in ("Pallet", "Vto", "Cajas"):
        tree4.column(col, anchor="center", width=70)
    elif col == "Producto":
        tree4.column(col, anchor="center", width=200)
    else:
        tree4.column(col, anchor="center", width=100)

# Empaquetar widgets
tree4.place(relx=0,rely=0.08, relheight=0.85,relwidth=0.96)
scrollbar_y.place(relx=0.95,rely=0.08, relheight=0.85,relwidth=0.05)
scrollbar_x.place(relx=0,rely=0.9, relheight=0.1,relwidth=0.98)
##################################################################################################
# Pestaña: Configuración
frame_config = ttk.Frame(notebook)
        
label_ruta = ttk.Label(frame_config, text="Ruta a Base de Datos")
label_ruta.place(relx=0.05, rely=0.7)        
entrada_ruta_bd = ttk.Entry(frame_config, width= 60)
entrada_ruta_bd.place(relx=0.27, rely=0.7)  
btn_guardar = ttk.Button(frame_config, text="Conf. Ruta", command = partial(selecionar_ruta,"bd"))
btn_guardar.place(relx=0.8, rely=0.7)

label_ruta_e = ttk.Label(frame_config, text="Ruta Exportar Planillas")
label_ruta_e.place(relx=0.05, rely=0.3)        
entrada_rutae = ttk.Entry(frame_config, width= 60)
entrada_rutae.place(relx=0.27, rely=0.3)  
btn_ruta_e = ttk.Button(frame_config, text="Conf. Ruta", command = partial(selecionar_ruta,"planilla"))
btn_ruta_e.place(relx=0.8, rely=0.3)


ttk.Button(frame_config, text="Buscar...", command = partial(selecionar_ruta,"excel")).place(relx=0.8, rely=0.5) 
entrada_ruta_excel = ttk.Entry(frame_config, width=30)
entrada_ruta_excel.place(relx=0.27, rely=0.5)
ttk.Label(frame_config, text="Plantilla Excel (.xlsx):").place(relx=0.05, rely=0.5)
#####################################################################################################
# Pestaña: OC
frame_oc = ttk.Frame(notebook)

contenedor_oc = ttk.Frame(frame_oc)
contenedor_oc.place(relx=0.01, rely=0.01, relwidth=0.6, relheight=0.5)  # Ajusta si es necesario
ttk.Label(contenedor_oc, text="Ordenes de Compra:").place(relx=0,rely=0)
# Scrollbars
scrollbar_y = ttk.Scrollbar(contenedor_oc, orient="vertical")
scrollbar_x = ttk.Scrollbar(contenedor_oc, orient="horizontal")

# Treeview con scrollbars
treeoc = ttk.Treeview(
    contenedor_oc,
    columns=("id","OC","Producto","Cajas","Producido","Pendiente","Destino","Cliente","Fecha Entrega","Fecha Despacho","Estado"), show='headings', 
    yscrollcommand=scrollbar_y.set,
    xscrollcommand=scrollbar_x.set,
    height=10
)

scrollbar_y.config(command=treeoc.yview)
scrollbar_x.config(command=treeoc.xview)

# Configurar columnas
for col in treeoc["columns"]:
    treeoc.heading(col, text=col)
    if col == "id":
        treeoc.column(col, width=0, stretch=False)
    elif col in ("Pallet", "Vto", "Cajas"):
        treeoc.column(col, anchor="center", width=70)
    elif col == "Producto":
        treeoc.column(col, anchor="center", width=200)
    else:
        treeoc.column(col, anchor="center", width=100)

treeoc.place(relx=0,rely=0.08, relheight=0.85,relwidth=0.96)
combo_var3 = tk.StringVar()
scrollbar_y.place(relx=0.95,rely=0.08, relheight=0.85,relwidth=0.05)
scrollbar_x.place(relx=0,rely=0.9, relheight=0.1,relwidth=0.98)
ttk.Label(frame_oc, text="Producto:").place(relx=0.62,rely=0.06)
producto_combo_oc = ttk.Combobox(frame_oc,width=50,textvariable=combo_var3)
producto_combo_oc.place(relx=0.67,rely=0.06)
producto_combo_oc.bind('<Return>', partial(filtrar_opciones,producto_combo_oc,"oc"))
ttk.Label(frame_oc, text="N° de OC:").place(relx=0.62,rely=0.11)
noc = ttk.Entry(frame_oc, width=30)
noc.place(relx=0.67,rely=0.11)
ttk.Label(frame_oc, text="Cantidad:").place(relx=0.62,rely=0.16,)
entry_cantidad = ttk.Entry(frame_oc, width=30,validate="key", validatecommand=vcmd)
entry_cantidad.place(relx=0.67,rely=0.16)
ttk.Label(frame_oc, text="Destino:").place(relx=0.62,rely=0.21)
combo_destino = ttk.Combobox(frame_oc,width=50)
combo_destino.place(relx=0.67,rely=0.21)
ttk.Label(frame_oc, text="Cliente:").place(relx=0.62,rely=0.26)
combo_cliente = ttk.Combobox(frame_oc,width=50)
combo_cliente.place(relx=0.67,rely=0.26)
ttk.Label(frame_oc, text="Fecha de Entrega:").place(relx=0.62,rely=0.31)
fecha_pedido = ttk.Entry(frame_oc, width=30)
fecha_pedido.place(relx=0.69,rely=0.31)

btn_cargar = ttk.Button(frame_oc, text="Agregar", command= agregar_oc)
btn_cargar.place(relx=0.69, rely=0.36)
btn_finalizar = ttk.Button(frame_oc, text="Finalizar", command= finalizar_oc)
btn_finalizar.place(relx=0.69, rely=0.42)

leer_archivo()
producto_combo['values'],combo_destino['values'],combo_cliente['values']  = leer_base()
producto_combo_oc['values'] = producto_combo['values']
producto_combo_sim['values']=producto_combo['values']
producto_combo.bind("<<ComboboxSelected>>",cargar_datos)

combo_destino['state'] = 'readonly'
combo_cliente['state'] = 'readonly'

notebook.add(frame_oc, text="Ordenes de Compra")
notebook.add(frame_config, text="Configuración")
root.mainloop()

